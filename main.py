import os
import cv2
import numpy as np
import time
from multiprocessing.pool import Pool
from openpyxl import Workbook


def main():
    directory = '2014-17-32-1'
    tp1 = time.time()

    images = [os.path.abspath(os.path.join(directory, filename))
              for filename in os.listdir(directory)
              if os.path.isfile(os.path.abspath(os.path.join(directory, filename)))]

    p = Pool(processes=4)
    results = p.map(process_image, images)
    p.close()
    p.join()

    tp2 = time.time()
    print(f'*** Общее время: {tp2 - tp1} ***')

    # Создание Excel-файла
    wb = Workbook()

    # Удаление стандартного листа, который создается по умолчанию
    wb.remove(wb.active)

    headers = ['Name', 'CountPeatles', 'SquarePeatles', 'DivisionHalfHorizontal', 'DivisionHalfVertical',
               'AverageRed', 'AverageGreen', 'AverageBlue',
               'AverageRedFirst', 'AverageGreenFirst', 'AverageBlueFirst',
               'AverageRedSecond', 'AverageGreenSecond', 'AverageBlueSecond',
               'AverageRedThird', 'AverageGreenThird', 'AverageBlueThird',
               'AverageRedFourth', 'AverageGreenFourth', 'AverageBlueFourth']

    for result in results:
        base_data = result[:2]  # название, количество объектов
        areas, division_half_horizontal, division_half_vertical, reds, greens, blues, avg_red_top, avg_green_top, avg_blue_top, avg_red_bottom, avg_green_bottom, avg_blue_bottom, avg_red_left, avg_green_left, avg_blue_left, avg_red_right, avg_green_right, avg_blue_right = \
            result[2], result[3], result[4], result[5], result[6], result[7], result[8], result[9], result[10], result[
                11], \
                result[12], result[13], result[14], result[15], result[16], result[17], result[18], result[
                19]  # площади объектов и средние значения

        # Создание нового листа
        sheet_name = os.path.basename(result[0])
        ws = wb.create_sheet(title=sheet_name)

        # Запись заголовков
        for col_num, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_num, value=header)

        # Запись основных данных
        ws.cell(row=2, column=1, value=base_data[0])
        ws.cell(row=2, column=2, value=base_data[1])

        # Запись площадей объектов в столбик
        for i, area in enumerate(areas, start=2):
            ws.cell(row=i, column=3, value=round(area))

        # Запись значений деленных линий в столбик
        for i, (half_horizontal, half_vertical) in enumerate(zip(division_half_horizontal, division_half_vertical),
                                                             start=2):
            ws.cell(row=i, column=4, value=round(half_horizontal))
            ws.cell(row=i, column=5, value=round(half_vertical))

        # Запись средних значений цветов в столбик
        for i, (red, green, blue) in enumerate(zip(reds, greens, blues), start=2):
            ws.cell(row=i, column=6, value=round(red))
            ws.cell(row=i, column=7, value=round(green))
            ws.cell(row=i, column=8, value=round(blue))

        # Запись средних значений цветов половин в столбик
        for i in range(len(avg_red_top)):
            ws.cell(row=i + 2, column=9, value=round(avg_red_top[i]))
            ws.cell(row=i + 2, column=10, value=round(avg_green_top[i]))
            ws.cell(row=i + 2, column=11, value=round(avg_blue_top[i]))
            ws.cell(row=i + 2, column=12, value=round(avg_red_bottom[i]))
            ws.cell(row=i + 2, column=13, value=round(avg_green_bottom[i]))
            ws.cell(row=i + 2, column=14, value=round(avg_blue_bottom[i]))
            ws.cell(row=i + 2, column=15, value=round(avg_red_left[i]))
            ws.cell(row=i + 2, column=16, value=round(avg_green_left[i]))
            ws.cell(row=i + 2, column=17, value=round(avg_blue_left[i]))
            ws.cell(row=i + 2, column=18, value=round(avg_red_right[i]))
            ws.cell(row=i + 2, column=19, value=round(avg_green_right[i]))
            ws.cell(row=i + 2, column=20, value=round(avg_blue_right[i]))
            print("Base Data:", base_data)
            print("Areas:", areas)
            print("Division Half Horizontal:", division_half_horizontal)
            print("Division Half Vertical:", division_half_vertical)
            print("Average Colors R:", reds)
            print("Average Colors G:", greens)
            print("Average Colors B:", blues)
            print("Avg Red Top:", avg_red_top)
            print("Avg Green Top:", avg_green_top)
            print("Avg Blue Top:", avg_blue_top)
            print("Avg Red Bottom:", avg_red_bottom)
            print("Avg Green Bottom:", avg_green_bottom)
            print("Avg Blue Bottom:", avg_blue_bottom)
            print("Avg Red Left:", avg_red_left)
            print("Avg Green Left:", avg_green_left)
            print("Avg Blue Left:", avg_blue_left)
            print("Avg Red Right:", avg_red_right)
            print("Avg Green Right:", avg_green_right)
            print("Avg Blue Right:", avg_blue_right)
    # Сохранение Excel-файла
    wb.save('2014-17-32oct22.xlsx')


def calculate_average_color(image, mask):
    # Преобразуем изображение и маску в одномерные массивы
    image = image.reshape((-1, 3))
    mask = mask.reshape((-1))

    # Оставляем только те пиксели, которые принадлежат лепестку
    petal_pixels = image[mask == 255]

    # Рассчитываем среднее значение цветовых каналов
    average_color = np.mean(petal_pixels, axis=0)

    return average_color


def calculate_max_lines(mask):
    # Применяем трансформацию расстояния
    dist_transform = cv2.distanceTransform(mask, cv2.DIST_L2, 5)

    # Находим координаты максимального значения (центра лепестка)
    _, _, _, max_loc = cv2.minMaxLoc(dist_transform)

    # Извлекаем маску лепестка
    petal_mask = mask == 255

    # Вычисляем максимальную горизонтальную линию
    horizontal_line = np.sum(petal_mask[max_loc[1], :])

    # Вычисляем максимальную вертикальную линию
    vertical_line = np.sum(petal_mask[:, max_loc[0]])

    return horizontal_line, vertical_line


def split_mask(mask, max_loc):
    # Делим маску на две половины по горизонтальной и вертикальной линиям
    h, w = mask.shape
    mask_top = mask.copy()
    mask_top[max_loc[1]:, :] = 0

    mask_bottom = mask.copy()
    mask_bottom[:max_loc[1], :] = 0

    mask_left = mask.copy()
    mask_left[:, max_loc[0]:] = 0

    mask_right = mask.copy()
    mask_right[:, :max_loc[0]] = 0

    return mask_top, mask_bottom, mask_left, mask_right


def process_image(path):
    print(f'*** Изображение: {path} ***')
    image = cv2.imread(path)
    b, g, r = cv2.split(image)

    # Используем cv2.inRange для сегментации по синему каналу
    mask = cv2.inRange(b, 5, 106)

    # Найдите контуры объектов
    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    # Подсчитайте количество объектов и их площади
    areas = [cv2.contourArea(contour) for contour in contours if cv2.contourArea(contour) > 20000]
    num_objects = len(areas)

    division_half_horizontal = []
    division_half_vertical = []
    average_colors_r = []
    average_colors_g = []
    average_colors_b = []
    avg_red_top = []
    avg_green_top = []
    avg_blue_top = []
    avg_red_bottom = []
    avg_green_bottom = []
    avg_blue_bottom = []
    avg_red_left = []
    avg_green_left = []
    avg_blue_left = []
    avg_red_right = []
    avg_green_right = []
    avg_blue_right = []

    for contour in contours:
        if cv2.contourArea(contour) > 15000:
            # Создаем маску для текущего лепестка
            petal_mask = np.zeros_like(mask)
            cv2.drawContours(petal_mask, [contour], -1, 255, thickness=cv2.FILLED)

            # Рассчитываем средний цвет лепестка
            average_color = calculate_average_color(image, petal_mask)
            average_colors_r.append(round(average_color[2]))
            average_colors_g.append(round(average_color[1]))
            average_colors_b.append(round(average_color[0]))

            # Вычисляем максимальные линии внутри лепестка
            max_horizontal, max_vertical = calculate_max_lines(petal_mask)

            # Находим координаты максимального значения (центра лепестка)
            _, _, _, max_loc = cv2.minMaxLoc(cv2.distanceTransform(petal_mask, cv2.DIST_L2, 3))

            # Делим маску лепестка на четыре части
            mask_top, mask_bottom, mask_left, mask_right = split_mask(petal_mask, max_loc)

            # Рассчитываем средние цвета для каждой половины
            avg_color_top = calculate_average_color(image, mask_top)
            avg_color_bottom = calculate_average_color(image, mask_bottom)
            avg_color_left = calculate_average_color(image, mask_left)
            avg_color_right = calculate_average_color(image, mask_right)

            # Добавляем средние цвета в соответствующие списки
            avg_red_top.append(avg_color_top[2])
            avg_green_top.append(avg_color_top[1])
            avg_blue_top.append(avg_color_top[0])
            avg_red_bottom.append(avg_color_bottom[2])
            avg_green_bottom.append(avg_color_bottom[1])
            avg_blue_bottom.append(avg_color_bottom[0])
            avg_red_left.append(avg_color_left[2])
            avg_green_left.append(avg_color_left[1])
            avg_blue_left.append(avg_color_left[0])
            avg_red_right.append(avg_color_right[2])
            avg_green_right.append(avg_color_right[1])
            avg_blue_right.append(avg_color_right[0])

            # Находим средние значения для горизонтальных и вертикальных половин

            division_half_horizontal.append(max_horizontal)
            division_half_vertical.append(max_vertical)

    return (path, num_objects, areas, division_half_horizontal, division_half_vertical, average_colors_r,
            average_colors_g, average_colors_b, avg_red_top, avg_green_top, avg_blue_top,
            avg_red_bottom, avg_green_bottom, avg_blue_bottom, avg_red_left, avg_green_left,
            avg_blue_left, avg_red_right, avg_green_right, avg_blue_right)


if __name__ == '__main__':
    main()
